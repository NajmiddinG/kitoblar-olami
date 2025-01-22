import sys
from PySide6.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QFileDialog, QMessageBox, QTableWidget, QToolTip 
from PySide6.QtCore import QTimer
from PySide6 import QtWidgets
from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
import create_table as db
from datetime import datetime, date
from PySide6.QtGui import QColor, QKeySequence, QShortcut, QAction, QIcon
from PySide6.QtCore import Qt, QDate
from ui import Ui_MainWindow

conn = db.conn
cur = db.cur

def tableResizeMode(table: QTableWidget):
    header = table.horizontalHeader()
    for i in range(header.count()):
        # table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)
        table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #333; /* Darker border around the entire table */
                gridline-color: #444; /* Darker color of the grid lines between cells */
                background-color: #222; /* Dark background color for the table */
            }
            QTableWidget::item {
                border: 1px solid #333; /* Darker border around each cell */
                background-color: #333; /* Dark background color for cells */
                color: #ddd; /* Light text color for cells */
            }
            QTableWidget::item:selected {
                background: #555; /* Darker background color of selected cells */
                border: 1px solid #777; /* Lighter border color of selected cells */
                color: #fff; /* Light text color for selected cells */
            }
            QTableWidget::item:hover {
                background: #444; /* Darker background color when hovering over a cell */
            }
            QHeaderView::section {
                background-color: #333; /* Dark background color of header sections */
                color: #ddd; /* Light text color of header sections */
                border: 1px solid #444; /* Border around header sections */
            }
            QHeaderView::indicator {
                width: 16px; /* Width of the indicator area (for sort arrows) */
                height: 16px; /* Height of the indicator area (for sort arrows) */
                border: none; /* No border for the indicator area */
                background-color: white; /* Background color for the indicator area */
            }
            QHeaderView::indicator:checked {
                background-color: white; /* Background color when checked (sorted) */
            }
        """)
        # header.setStyleSheet("color: black;")  

def tableResizeMode2(table: QTableWidget):
    header = table.horizontalHeader()
    for i in range(header.count()):
        # table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(i, QtWidgets.QHeaderView.ResizeToContents)
        table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #333; /* Dark border around the entire table */
                gridline-color: #444; /* Darker color for the grid lines */
                background-color: #2b3e3a; /* Darker teal background for the table */
            }
            QTableWidget::item {
                border: 1px solid #333; /* Darker border around each cell */
                background-color: #1f2f2b; /* Darker teal for cells */
                color: #ddd; /* Light text color for cells */
            }
            QTableWidget::item:selected {
                background: #33675f; /* Slightly lighter teal background when selected */
                border: 1px solid #555; /* Lighter border color when selected */
                color: #fff; /* White text for selected cells */
            }
            QTableWidget::item:hover {
                background: #3d5c56; /* Hover effect with a slightly lighter background */
            }
            QHeaderView::section {
                background-color: #1c2c28; /* Darker teal background for header sections */
                color: #ddd; /* Light text color for header sections */
                border: 1px solid #444; /* Border around header sections */
            }
            QHeaderView::indicator {
                width: 16px; /* Width of the indicator area (for sort arrows) */
                height: 16px; /* Height of the indicator area (for sort arrows) */
                border: none; /* No border for the indicator area */
                background-color: #ddd; /* Light background color for the indicator area */
            }
            QHeaderView::indicator:checked {
                background-color: #ddd; /* Light background color when sorted */
            }
        """)
        # header.setStyleSheet("color: black;")  

class DiamondWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(DiamondWindow, self).__init__()
        self.setupUi(self)
        self.table_resize()
        self.initialize_variables()
        self.connect_handlers()

    def table_resize(self):
        tableResizeMode(self.tableWidget)
        tableResizeMode(self.tableWidget_2)
        tableResizeMode(self.tableWidget_3)
        tableResizeMode(self.tableWidget_4)
        tableResizeMode(self.tableWidget_5)
        tableResizeMode2(self.tableWidget_6)
        tableResizeMode2(self.tableWidget_7)
        tableResizeMode(self.tableWidget_8)
        tableResizeMode(self.tableWidget_9)

    def initialize_variables(self):
        self.setGeometry(0, 0, 2400, 1000)
        self.addproductwidget = None
        self.tableWidget_2.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.sotuv = {} # Dictionary tartib name: (List(book id, number))
        self.curr_sotuv = "Sotuv 1"
        self.prev_sotuv = 0
        self.setCentralWidget(self.tabWidget)
        self.handle_tabbar_clicked(0)
        self.tabWidget.tabBarClicked.connect(self.handle_tabbar_clicked, Qt.UniqueConnection)
        self.label_change()
        qmenu = self.menuBar()
        qmenu.setStyleSheet("""
        QMenuBar::item:selected {
            background-color: blue;
            color: white;
        }
        QMenuBar::item {
            border: 1px solid rgb(94, 111, 135);
            padding: 5px;
        }
        """)

        self.label_16.setStyleSheet("""
        QLabel {
            color: #9CA7B7;
            background-color: #080A0C;
            border-top-right-radius: 5px;
            border-bottom-right-radius: 5px;
            padding-top: 5px;
            padding-right: 10px;
            padding-bottom: 5px;
            padding-left: 10px;
            border: 1px solid rgb(94, 111, 135);
            border-left: 0px
        }
        """)
        self.label_17.setStyleSheet("""
            QLabel {
                color: #ddd; /* Light text color for better contrast */
                background-color: #1f2f2b; /* Dark teal background, complementary to page background */
                border-top-right-radius: 5px;
                border-bottom-right-radius: 5px;
                padding-top: 5px;
                padding-right: 10px;
                padding-bottom: 5px;
                padding-left: 10px;
                border: 1px solid #33675f; /* A slightly lighter teal border */
                border-left: 0px; /* Removing the left border for a cleaner design */
            }
        """)

        line_edit_style = """
        QLineEdit {
            background-color: #333; /* Dark background color for the line edit */
            color: #ddd; /* Light text color for the line edit */
            border: 1px solid rgb(94, 111, 135);
            padding: 5px; /* Padding inside the line edit */
            border-radius: 3px; /* Rounded corners for the border */
        }
        QLineEdit::placeholder {
            color: #888; /* Color of the placeholder text */
        }
        """
        self.lineEdit_2.setStyleSheet(line_edit_style)
        self.lineEdit_3.setStyleSheet(line_edit_style)
        self.lineEdit_4.setStyleSheet(line_edit_style)
        self.lineEdit_6.setStyleSheet(line_edit_style)
        self.lineEdit_7.setStyleSheet(line_edit_style)
        customline_edit_style = """
        QLineEdit{
            border-top-left-radius: 5px;
            border-bottom-left-radius: 5px;
            border: 1px solid rgb(94, 111, 135);
            padding: 5px;
            padding-left: 15px;
            border-right: 0px;
        }
        QLineEdit::placeholder {
            color: #888; /* Color of the placeholder text */
        }
        """
        self.lineEdit.setStyleSheet(customline_edit_style)
        self.lineEdit_5.setStyleSheet(customline_edit_style)
        
        combo_box_style = """
        QComboBox {
            border: 1px solid rgb(94, 111, 135);
            height: 35px;
            padding-left: 8px;
            background-color: #333; /* Dark background for the combo box */
            color: #ddd; /* Light text color for the combo box */
            border-radius: 3px; /* Rounded corners for the border */
        }
        QComboBox QAbstractItemView {
            background-color: #333; /* Dark background for the drop-down list */
            color: #ddd; /* Light text color for items in the drop-down list */
            selection-background-color: #555; /* Background color for selected items */
            selection-color: #fff; /* Text color for selected items */
        }
        QComboBox::item:selected {
            background-color: #555; /* Background color for selected item in the drop-down list */
            color: #fff; /* Text color for selected item in the drop-down list */
        }
        """
        self.comboBox.setStyleSheet(combo_box_style)
        self.comboBox_2.setStyleSheet(combo_box_style)
        self.comboBox_3.setStyleSheet(combo_box_style)
        self.comboBox_4.setStyleSheet(combo_box_style)

        success_button_style = """
            QPushButton {
                background-color: rgb(31, 118, 28); /* Dark green background for success */
                color: white; /* White text color */
                border: 1px solid rgb(94, 111, 135);
                border-radius: 5px; /* Rounded corners */
                padding-left: 5px; /* Padding inside the button */
                padding-right: 5px; /* Padding inside the button */
                padding-top: 2px; /* Padding inside the button */
                padding-bottom: 2px; /* Padding inside the button */
                font-size: 20px; /* Font size for the text */
                max-height: 32px; /* Maximum height of the button */
                min-height: 10px; /* Maximum height of the button */
                min-width: 20px; /* Maximum height of the button */
            }
            QPushButton:hover {
                background-color: rgb(24, 87, 21); /* Slightly darker green on hover */
                opacity: 0.9; /* Slightly transparent on hover */
            }
            QPushButton:pressed {
                background-color: rgb(18, 65, 16); /* Even darker green when pressed */
                opacity: 0.8; /* More transparent when pressed */
            }
        """
        danger_button_style = """
            QPushButton {
                background-color: rgb(160, 20, 20); /* Dark red background for danger */
                color: white; /* White text color */
                border: 1px solid rgb(94, 111, 135);
                border-radius: 5px; /* Rounded corners */
                padding-left: 5px; /* Padding inside the button */
                padding-right: 5px; /* Padding inside the button */
                padding-top: 2px; /* Padding inside the button */
                padding-bottom: 2px; /* Padding inside the button */
                font-size: 20px; /* Font size for the text */
                max-height: 32px; /* Maximum height of the button */
                min-height: 10px; /* Maximum height of the button */
                min-width: 20px; /* Maximum height of the button */
            }
            QPushButton:hover {
                background-color: rgb(140, 20, 20); /* Slightly darker red on hover */
                opacity: 0.9; /* Slightly transparent on hover */
            }
            QPushButton:pressed {
                background-color: rgb(120, 20, 20); /* Even darker red when pressed */
                opacity: 0.8; /* More transparent when pressed */
            }
        """
        self.pushButton.setStyleSheet(danger_button_style)
        self.pushButton_8.setStyleSheet(danger_button_style)
        self.pushButton_9.setStyleSheet(danger_button_style)
        self.pushButton_2.setStyleSheet(success_button_style)
        self.pushButton_3.setStyleSheet(success_button_style)
        self.pushButton_4.setStyleSheet(success_button_style)
        self.pushButton_5.setStyleSheet(success_button_style)
        self.pushButton_14.setStyleSheet(danger_button_style)
        self.pushButton_6.setStyleSheet(success_button_style)
        self.pushButton_7.setStyleSheet(success_button_style)
        self.pushButton_10.setStyleSheet(success_button_style)
        self.pushButton_11.setStyleSheet(success_button_style)
        self.pushButton_12.setStyleSheet(success_button_style)
        self.pushButton_13.setStyleSheet(success_button_style)
        self.pushButton_15.setStyleSheet(success_button_style)
        self.pushButton_16.setStyleSheet(danger_button_style)
        self.pushButton_17.setStyleSheet(danger_button_style)

        # menus
        menubar = self.menuBar()
        file_menu = menubar.addMenu('Funksiyalar')

        # Creating QAction for each item
        qidirish = QAction('Qidirish                  F3', self)
        sotish = QAction('Sotish                     F4', self)
        # big = QAction('Kattalashtirish            F5', self)
        # small = QAction('Kichiklashtirish           F6', self)

        # Connecting triggered signal of each QAction to a function
        qidirish.triggered.connect(self.Key_F3_function)
        sotish.triggered.connect(self.F4_func_control)
        # big.triggered.connect(self.make_big)
        # small.triggered.connect(self.make_small)

        # Adding actions to the menu
        file_menu.addAction(qidirish)
        file_menu.addAction(sotish)
        # file_menu.addAction(big)
        # file_menu.addAction(small)



        shortcut = QKeySequence(Qt.Key_F3)
        self.shortcut = QShortcut(shortcut, self)
        self.shortcut.activated.connect(self.Key_F3_function)

        shortcut = QKeySequence(Qt.Key_F4)
        self.shortcut = QShortcut(shortcut, self)
        self.shortcut.activated.connect(self.F4_func_control)
        self.Key_F3_function()

    def connect_handlers(self):
        self.lineEdit_2.textChanged.connect(self.on_line_edit_changed, Qt.UniqueConnection)
        self.lineEdit_2.returnPressed.connect(self.scanner_returned, Qt.UniqueConnection)
        self.tableWidget_2.cellDoubleClicked.connect(self.add_selected_item_to_table, Qt.UniqueConnection)
        self.pushButton_5.clicked.connect(self.add_selected_item_to_table, Qt.UniqueConnection)
        self.pushButton_14.clicked.connect(self.subtract_selected_item_to_table, Qt.UniqueConnection)
        self.pushButton_10.clicked.connect(self.add_selected_item_to_new_table, Qt.UniqueConnection)
        self.pushButton_2.clicked.connect(self.accept_buy, Qt.UniqueConnection)
        self.pushButton.clicked.connect(self.cancel_buy)
        self.tableWidget.itemChanged.connect(self.handle_item_changed, Qt.UniqueConnection)
        self.tableWidget.clicked.connect(self.clicked_cancel, Qt.UniqueConnection)
        self.comboBox_2.currentIndexChanged.connect(self.sell_combo_control, Qt.UniqueConnection)
        self.comboBox_3.currentIndexChanged.connect(self.filter_history_combo, Qt.UniqueConnection)
        self.comboBox_4.currentIndexChanged.connect(self.filter_history2_combo, Qt.UniqueConnection)
        self.pushButton_8.clicked.connect(self.sotuv_table_remove, Qt.UniqueConnection)
        self.pushButton_3.clicked.connect(self.save_tableWidget_data_to_database, Qt.UniqueConnection)
        self.pushButton_13.clicked.connect(self.create_order, Qt.UniqueConnection)
        self.pushButton_6.clicked.connect(self.clicked_export_book, Qt.UniqueConnection)
        self.pushButton_4.clicked.connect(self.clicked_new_book, Qt.UniqueConnection)
        self.lineEdit_3.textChanged.connect(self.on_line_edit_changed, Qt.UniqueConnection)
        self.tableWidget_5.itemSelectionChanged.connect(self.handle_tableWidget_5_selected, Qt.UniqueConnection)
        self.pushButton_7.clicked.connect(self.filter_history, Qt.UniqueConnection)
        self.lineEdit_4.textChanged.connect(self.on_line_edit_changed2, Qt.UniqueConnection)
        self.lineEdit_4.returnPressed.connect(self.scanner_returned2, Qt.UniqueConnection)
        self.tableWidget_6.cellDoubleClicked.connect(self.add_selected_item_to_table2, Qt.UniqueConnection)
        self.pushButton_12.clicked.connect(self.add_selected_item_to_table2, Qt.UniqueConnection)
        self.pushButton_16.clicked.connect(self.subtract_selected_item_to_table2, Qt.UniqueConnection)
        self.pushButton_11.clicked.connect(self.accept_buy2, Qt.UniqueConnection)
        self.pushButton_9.clicked.connect(self.cancel_buy2, Qt.UniqueConnection)
        self.tableWidget_7.itemChanged.connect(self.handle_item_changed2, Qt.UniqueConnection)
        self.tableWidget_7.clicked.connect(self.clicked_cancel2, Qt.UniqueConnection)
        self.tableWidget_8.itemSelectionChanged.connect(self.handle_tableWidget_8_selected, Qt.UniqueConnection)
        self.pushButton_15.clicked.connect(self.filter_history2, Qt.UniqueConnection)
        self.pushButton_17.clicked.connect(self.return_product, Qt.UniqueConnection)

    def Key_F3_function(self):
        index = self.tabWidget.currentIndex()
        if index==0:
            self.lineEdit_2.setFocus()
        elif index==4:
            self.lineEdit_4.setFocus()
        elif index==1:
            self.lineEdit_3.setFocus()

    def F4_func_control(self):
        index = self.tabWidget.currentIndex()
        if index==0:
            self.accept_buy()
        elif index==4:
            pass
            # self.accept_buy2()
        elif index==1:
            pass
            # self.save_tableWidget_data_to_database()
        self.Key_F3_function()

    def Key_F3_2_function(self):
        self.lineEdit_3.setFocus()

    def label_change(self):
        try:
            text = self.label_3.text()
            index = text.index('Kitob')
            text = text[:index]+'📒 '+text[index:]
            self.label_3.setText(text)

            text = self.label_2.text()
            index = text.index('Umumiy')
            text = text[:index]+'💰 '+text[index:]
            self.label_2.setText(text)
        except:
            pass
        self.Key_F3_function()

    def handle_tabbar_clicked(self, index):
        if index==0: self.sotuv_tab()
        if index==4: self.tovar_tab()
        elif index==1: self.mahsulot_tab()
        elif index==2: self.sotuv_tarix_tab()
        elif index==3: self.tovar_tarix_tab()
        style_sheet = """
        QTabWidget::pane {
            border: 1px solid rgb(94, 111, 135);
            border-radius: 5px; /* Rounded corners for the tab widget pane */
            background-color: #222; /* Dark background for the tab pane */
        }
        QTabBar::tab {
            background-color: #333; /* Dark background for tabs */
            color: #ddd; /* Light text color for tabs */
            border: 1px solid rgb(94, 111, 135);
            border-radius: 5px; /* Rounded corners for tabs */
            padding: 10px; /* Padding inside the tabs */
            min-width: 80px; /* Minimum width of tabs */
        }
        QTabBar::tab:selected {
            background-color: #444; /* Darker background for the selected tab */
            color: #fff; /* Light text color for the selected tab */
        }
        QTabBar::tab:hover {
            background-color: #555; /* Slightly lighter background on hover */
        }
        QTabWidget::tab-bar {
            alignment: center; /* Center-align the tabs */
        }
        """
        self.tabWidget.setStyleSheet(style_sheet)
        self.tabWidget.setCurrentIndex(index)
        self.Key_F3_function()

    def sotuv_tab(self):
        self.Key_F3_function()

        # Timer for the delay
        self.timer = QTimer(self)
        self.timer.setSingleShot(True)
        self.timer.timeout.connect(self.search_database)
        self.Key_F3_function()

    def scanner_returned(self):
        scanned_text = self.lineEdit_2.text()
        self.Key_F3_function()

    def clicked_cancel(self, item):
        print("clicked")
        if item.column() == 7:
            if 0 <= item.row() < self.tableWidget.rowCount():
                self.tableWidget.removeRow(item.row())
            else:
                pass
            self.calculate()
            self.Key_F3_function()

    def handle_item_changed(self, item):
        # Check if the item is in column 2
        if item.column() == 2:
            try: self.calculate()
            except: self.cancel_buy()
        self.Key_F3_function()

    def calculate(self):
        column_index = 2
        total_amount = 0

        for row_index in range(self.tableWidget.rowCount()):
            item = self.tableWidget.item(row_index, column_index)

            if item is not None:
                count = int(float(item.text()))
                id = int(float(self.tableWidget.item(row_index, 0).text()))
                cur.execute("SELECT narxi FROM Kitob WHERE id=?", (id,))
                narxi = int(float(cur.fetchone()[0]))
                cur.execute("SELECT qoldiq FROM Kitob WHERE id=?", (id,))
                qoldiq = int(float(cur.fetchone()[0]))
                if count>qoldiq:
                    self.tableWidget.setItem(row_index, 2, QTableWidgetItem(str(qoldiq)))
                    count = qoldiq
                total_amount += narxi * count
                self.tableWidget.setItem(row_index, 6, QTableWidgetItem(str(narxi*count)))

        self.lineEdit.setText(self.spacecomma(int(total_amount)))
        self.Key_F3_function()

    def cancel_buy(self):
        self.tableWidget_2.setRowCount(0)
        self.tableWidget.setRowCount(0)
        self.lineEdit_2.clear()
        self.lineEdit.setText("")
        self.lineEdit.setText("")
        self.lineEdit_7.setText("")
        self.comboBox.setCurrentIndex(0)
        self.Key_F3_function()

    def accept_buy(self):
        try:
            rows = self.tableWidget.rowCount()
            if rows:
                current_datetime = datetime.now()
                formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
                tex  = self.lineEdit.text()
                umumiy_hisob = ''
                kimga = self.lineEdit_7.text()
                if not kimga: kimga = "Nomalum"
                for i in tex:
                    if i.isdigit():
                        umumiy_hisob+=i
                cur.execute("INSERT INTO Tarix (kimga, sana, hisob, tolov_turi) VALUES (?, ?, ?, ?)", (kimga, formatted_datetime, umumiy_hisob, self.comboBox.currentText()))
                conn.commit()
                tarix_id = cur.lastrowid
                for row_index in range(rows):
                    kitob_id = int(float(self.tableWidget.item(row_index, 0).text()))
                    soni = int(self.tableWidget.item(row_index, 2).text())
                    qoldiq = int(self.tableWidget.item(row_index, 5).text())
                    hisob = int(self.tableWidget.item(row_index, 6).text())
                    cur.execute("UPDATE Kitob SET qoldiq = ? WHERE id = ?", (qoldiq-soni, kitob_id))
                    conn.commit()
                    cur.execute("INSERT INTO TarixItem (Tarix, Kitob, soni, hisob) VALUES (?, ?, ?, ?)", (tarix_id, kitob_id, soni, hisob))
                    conn.commit()
            QTimer.singleShot(1000, self.close_message_box)
            self.message_info = "Sotildi"
            self.message_type = 1
            QTimer.singleShot(50, self.show_message)
        except:
            QTimer.singleShot(1000, self.close_message_box)
            self.message_info = "Xatolik"
            self.message_type = 0
            QTimer.singleShot(50, self.show_message)
        self.cancel_buy()
        self.Key_F3_function()

    def return_product(self):
        row_index = self.tableWidget_5.currentRow()
        if row_index < 0: 
            return

        tarix_id = self.tableWidget_5.item(row_index, 0).text()

        # Create a message box to confirm the action
        confirmation = QMessageBox.question(
            self, 
            'Vazvratni tasdiqlash', 
            f"#'{tarix_id}' ni vazvrat qilmoqchimiz?", 
            QMessageBox.Yes | QMessageBox.No
        )

        # If user selects 'Yes', proceed
        if confirmation == QMessageBox.Yes:
            try:
                cur.execute("SELECT Kitob, soni FROM TarixItem WHERE Tarix = ?", (tarix_id,))
                tarix_items = cur.fetchall()  # [(Kitob, soni), ...]

                for kitob_id, soni in tarix_items:
                    cur.execute("SELECT qoldiq FROM Kitob WHERE id = ?", (kitob_id,))
                    current_stock = cur.fetchone()[0]
                    new_stock = int(current_stock) + int(soni)
                    cur.execute("UPDATE Kitob SET qoldiq = ? WHERE id = ?", (new_stock, kitob_id))

                cur.execute("DELETE FROM TarixItem WHERE Tarix = ?", (tarix_id,))
                cur.execute("DELETE FROM Tarix WHERE id = ?", (tarix_id,))
                conn.commit()

                print(f"Return completed for Tarix ID {tarix_id}. Stock updated and records deleted.")

            except Exception as e:
                conn.rollback()
                print(f"Error occurred: {e}")

        self.filter_history()

    def on_line_edit_changed(self):
        # Start the timer when the text in the line edit changes
        self.timer.start(1000)  # 1000 milliseconds = 1 second
        self.Key_F3_function()

    def search_database(self):
        # Perform the database query based on user input
        user_input = self.lineEdit_2.text()
        # Clear the table before populating it with new data
        # self.tableWidget_2.setRowCount(0)
        if user_input:
            if len(user_input)>7 and user_input.isdigit():
                query = "SELECT id, nomi, narxi, qoldiq, barcode FROM Kitob WHERE (barcode = ?) AND qoldiq > 0"
                cur.execute(query, (user_input,))
                result = cur.fetchall()
                if len(result)==1:
                    self.display_data_in_table(result, self.tableWidget_2)
                    self.tableWidget_2.selectRow(0)
                    self.add_selected_item_to_table()
                elif len(result)>1:
                    self.display_data_in_table(result, self.tableWidget_2)
                    self.tableWidget_2.selectRow(0)
                else:
                    self.lineEdit_2.clear()
                    QTimer.singleShot(1000, self.close_message_box)
                    self.message_info = "Topilmadi"
                    self.message_type = 0
                    QTimer.singleShot(50, self.show_message)
            else:
                query = "SELECT id, nomi, narxi, qoldiq, barcode FROM Kitob WHERE (nomi LIKE ? OR id LIKE ?) AND qoldiq > 0"
                cur.execute(query, (f'%{user_input}%', f'%{user_input}%'))
                result = cur.fetchall()
                if len(result):
                    self.display_data_in_table(result, self.tableWidget_2)
                    self.tableWidget_2.selectRow(0)
        self.Key_F3_function()

    def sotuv_table_remove(self):
        try:
            curr_combo = self.comboBox_2.currentIndex()
            curr_combo_text = self.comboBox_2.currentText()
            if self.comboBox_2.count()!=1:

                try: del self.sotuv[curr_combo_text]
                except: pass
                a = -1
                self.comboBox_2.clear()
                for key in self.sotuv.keys():
                    if self.sotuv[key]:
                        self.comboBox_2.addItem(key)
                        a+=1
                if a==-1:
                    self.comboBox_2.addItem('Sotuv 1')
                    self.curr_sotuv = 'Sotuv 1'
                else:
                    self.comboBox_2.setCurrentIndex(a)
                    self.curr_sotuv = self.comboBox_2.currentText()
        except: pass
        self.tableWidget.setRowCount(0)
        try:
            curr_sell = self.sotuv[self.comboBox_2.currentText()]
        except: return
        for id, number in curr_sell:
            check = self.check_tablewidget_add(id)
            if check==-1: 
                row_position = self.tableWidget.rowCount()
                self.tableWidget.insertRow(row_position)
                cur.execute("SELECT * FROM Kitob WHERE id=?", (id,))
                data = cur.fetchone()
                if int(data[4]):
                    for col_num, col_data in enumerate(data):
                        item = QTableWidgetItem(str(col_data))
                        if col_num>1: col_num+=1
                        self.tableWidget.setItem(row_position, col_num, item)
                    self.tableWidget.setItem(row_position, 2, QTableWidgetItem(str(number)))
                    column_position = 7
                    item = QTableWidgetItem('Bekor qilish')
                    background_color = QColor(160, 20, 20)
                    color = QColor(255, 255, 255)
                    item.setBackground(background_color)
                    item.setForeground(color)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.tableWidget.setItem(row_position, column_position, item)
                    item = self.tableWidget.item(row_position, 2)
                    item.setBackground(QColor(0, 0, 255))
                    item.setForeground(QColor(255, 255, 255))
                    item = self.tableWidget.item(row_position, 3)
                    item.setBackground(QColor(60, 179, 113))
                    item.setForeground(QColor(255, 255, 255))

                else:
                    current_row_count = self.tableWidget.rowCount()
                    if current_row_count > 0:
                        self.tableWidget.removeRow(current_row_count - 1)
            else:
                cur.execute("SELECT * FROM Kitob WHERE id=?", (id,))
                data = cur.fetchone()
                qoldiq = int(float(data[4]))
                prev = int(float(self.tableWidget.item(check, 2).text()))
                if prev<qoldiq:
                    self.tableWidget.setItem(check, 2, QTableWidgetItem(str(prev+1)))
                    item = self.tableWidget.item(check, 2)
                    item.setBackground(QColor(0, 0, 255))
                    item.setForeground(QColor(255, 255, 255))
        # self.tableWidget_2.setRowCount(0)
        self.curr_sotuv = self.comboBox_2.currentText()
        self.Key_F3_function()

    def sell_combo_control(self):
        self.sotuv[self.curr_sotuv] = []
        for index in range(self.tableWidget.rowCount()):
            self.sotuv[self.curr_sotuv].append([int(self.tableWidget.item(index, 0).text()), int(float(self.tableWidget.item(index, 2).text()))])
            # self.prev_sotuv=len(self.sotuv)
        count = self.comboBox_2.count()
        self.tableWidget.setRowCount(0)
        try:
            curr_sell = self.sotuv[self.comboBox_2.currentText()]
        except: return
        for id, number in curr_sell:
            check = self.check_tablewidget_add(id)
            if check==-1: 
                row_position = self.tableWidget.rowCount()
                self.tableWidget.insertRow(row_position)
                cur.execute("SELECT * FROM Kitob WHERE id=?", (id,))
                data = cur.fetchone()
                if int(float(data[4])):
                    for col_num, col_data in enumerate(data):
                        item = QTableWidgetItem(str(col_data))
                        if col_num>1: col_num+=1
                        self.tableWidget.setItem(row_position, col_num, item)
                    self.tableWidget.setItem(row_position, 2, QTableWidgetItem(str(number)))
                    column_position = 7
                    item = QTableWidgetItem('Bekor qilish')
                    background_color = QColor(160, 20, 20)
                    color = QColor(255, 255, 255)
                    item.setBackground(background_color)
                    item.setForeground(color)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.tableWidget.setItem(row_position, column_position, item)
                    item = self.tableWidget.item(row_position, 2)
                    item.setBackground(QColor(0, 0, 255))
                    item.setForeground(QColor(255, 255, 255))
                    item = self.tableWidget.item(row_position, 3)
                    item.setBackground(QColor(60, 179, 113))
                    item.setForeground(QColor(255, 255, 255))
                else:
                    current_row_count = self.tableWidget.rowCount()
                    if current_row_count > 0:
                        self.tableWidget.removeRow(current_row_count - 1)
            else:
                cur.execute("SELECT * FROM Kitob WHERE id=?", (id,))
                data = cur.fetchone()
                qoldiq = int(float(data[4]))
                prev = int(float(self.tableWidget.item(check, 2).text()))
                if prev<qoldiq:
                    self.tableWidget.setItem(check, 2, QTableWidgetItem(str(prev+1)))
                    item = self.tableWidget.item(check, 2)
                    item.setBackground(QColor(0, 0, 255), 1)
                    item.setForeground(QColor(255, 255, 255))
        # self.tableWidget_2.setRowCount(0)
        self.curr_sotuv = self.comboBox_2.currentText()
        self.Key_F3_function()

    def add_selected_item_to_new_table(self):
        current_tartib = self.comboBox_2.currentText()
        self.sotuv[current_tartib] = []
        for index in range(self.tableWidget.rowCount()):
            self.sotuv[current_tartib].append([int(self.tableWidget.item(index, 0).text()), int(float(self.tableWidget.item(index, 2).text()))])
        count = self.comboBox_2.count()
        curr_combo_text = self.comboBox_2.itemText(count - 1).split()[1]
        new_item_text = f'Sotuv {int(curr_combo_text) + 1}'
        self.comboBox_2.addItem(new_item_text)

        # Find the index of the newly added item
        new_item_index = self.comboBox_2.findText(new_item_text)

        self.comboBox_2.setCurrentIndex(new_item_index)
        
        self.curr_sotuv = f'Sotuv {int(curr_combo_text)+1}'
        self.tableWidget.setRowCount(0)
        # self.add_selected_item_to_table()
        self.Key_F3_function()

    def add_selected_item_to_table(self):
        selected_item = self.tableWidget_2.currentRow()
        self.lineEdit_2.clear()
        if selected_item>-1:
            id = int(float(self.tableWidget_2.item(selected_item, 0).text()))
            check = self.check_tablewidget_add(id)
            if check==-1: 
                row_position = self.tableWidget.rowCount()
                self.tableWidget.insertRow(row_position)
                cur.execute("SELECT * FROM Kitob WHERE id=?", (id,))
                data = cur.fetchone()
                if int(float(data[4])):
                    for col_num, col_data in enumerate(data):
                        item = QTableWidgetItem(str(col_data))
                        if col_num>1: col_num+=1
                        self.tableWidget.setItem(row_position, col_num, item)
                    self.tableWidget.setItem(row_position, 2, QTableWidgetItem('1'))
                    column_position = 7
                    item = QTableWidgetItem('Bekor qilish')
                    background_color = QColor(160, 20, 20)
                    color = QColor(255, 255, 255)
                    item.setBackground(background_color)
                    item.setForeground(color)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.tableWidget.setItem(row_position, column_position, item)
                    item = self.tableWidget.item(row_position, 2)
                    item.setBackground(QColor(0, 0, 255))
                    item.setForeground(QColor(255, 255, 255))
                    item = self.tableWidget.item(row_position, 3)
                    item.setBackground(QColor(60, 179, 113))
                    item.setForeground(QColor(255, 255, 255))
                    # for row in range(self.tableWidget.rowCount()-1, 0, -1):
                    #     for col in range(self.tableWidget.columnCount()):
                    #         new_item = self.tableWidget.item(row, col)
                    #         prev = self.tableWidget.item(row-1, col)
                    #         a1, a2 = str(new_item.text()), str(prev.text())
                    #         print(a1, a2)
                    #         new_item.setText(str(a2))
                    #         prev.setText(str(a1))
                    # print("---------------------------------------------")
                else:
                    current_row_count = self.tableWidget.rowCount()
                    if current_row_count > 0:
                        self.tableWidget.removeRow(current_row_count - 1)
            else:
                cur.execute("SELECT * FROM Kitob WHERE id=?", (id,))
                data = cur.fetchone()
                qoldiq = int(float(data[4]))
                prev = int(float(self.tableWidget.item(check, 2).text()))
                if prev<qoldiq:
                    self.tableWidget.setItem(check, 2, QTableWidgetItem(str(prev+1)))
                    item = self.tableWidget.item(check, 2)
                    item.setBackground(QColor(0, 0, 255))
                    item.setForeground(QColor(255, 255, 255))
        # self.tableWidget_2.setRowCount(0)
        self.Key_F3_function()
    
    def subtract_selected_item_to_table(self):
        selected_item = self.tableWidget_2.currentRow()
        # self.lineEdit_2.clear()
        if selected_item>-1:
            id = int(float(self.tableWidget_2.item(selected_item, 0).text()))
            check = self.check_tablewidget_add(id)
            if check!=-1:
                cur.execute("SELECT * FROM Kitob WHERE id=?", (id,))
                data = cur.fetchone()
                qoldiq = int(float(data[4]))
                prev = int(float(self.tableWidget.item(check, 2).text()))
                if 0<prev:
                    self.tableWidget.setItem(check, 2, QTableWidgetItem(str(prev-1)))
                    item = self.tableWidget.item(check, 2)
                    item.setBackground(QColor(0, 0, 255))
                    item.setForeground(QColor(255, 255, 255))
        # self.tableWidget_2.setRowCount(0)
        self.Key_F3_function()

    def check_tablewidget_add(self, id):
        is_present = -1
        for row_index in range(self.tableWidget.rowCount()):
            if self.tableWidget.item(row_index, 0).text() == str(id):
                is_present=row_index
                break
        return is_present
    
    def mahsulot_tab(self):
        self.populate_tableWidget_4()

        self.timer = QTimer(self)
        self.timer.setSingleShot(True)
        self.timer.timeout.connect(self.search_database_2)
        self.Key_F3_2_function()
        self.Key_F3_function()
    
    def search_database_2(self):
        user_input = self.lineEdit_3.text()
        if user_input:
            self.tableWidget_4.setRowCount(0)
            self.populate_tableWidget_4(filter=user_input)
        else: self.populate_tableWidget_4()
        self.Key_F3_2_function()
        self.Key_F3_function()

    def create_order(self):
        try:
            query = "SELECT id, nomi, barcode, tan_narx, kelgan_sana, kimdan FROM Kitob WHERE buyurtma != '-1' AND CAST(qoldiq AS INTEGER) <= CAST(buyurtma AS INTEGER)"
            df = pd.read_sql_query(query, conn)

            file_path, _ = QFileDialog.getSaveFileName(self, 'Save Excel File', '', 'Excel Files (*.xlsx)')

            if file_path:
                # Create a workbook and add a worksheet
                workbook = Workbook()
                worksheet = workbook.active

                # Customize the header row
                header_row = df.columns
                for col_num, value in enumerate(header_row, 1):
                    cell = worksheet.cell(row=1, column=col_num, value=value.title() if type(value)==str else value)
                    cell.font = Font(size=14, bold=True)

                for row_num, (_, row) in enumerate(df.iterrows(), 2):
                    for col_num, value in enumerate(row, 1):
                        worksheet.cell(row=row_num, column=col_num, value=value)

                # Save the workbook to the specified file path
                workbook.save(file_path)
        except: pass
        self.Key_F3_2_function()
        self.Key_F3_function()

    def clicked_export_book(self):
        try:
            query = "SELECT * FROM Kitob"
            df = pd.read_sql_query(query, conn)

            file_path, _ = QFileDialog.getSaveFileName(self, 'Save Excel File', '', 'Excel Files (*.xlsx)')

            if file_path:
                # Create a workbook and add a worksheet
                workbook = Workbook()
                worksheet = workbook.active

                # Customize the header row
                header_row = df.columns
                for col_num, value in enumerate(header_row, 1):
                    cell = worksheet.cell(row=1, column=col_num, value=value.title() if type(value)==str else value)
                    cell.font = Font(size=20, bold=True)

                # Add data to the worksheet
                for row_num, (_, row) in enumerate(df.iterrows(), 2):
                    for col_num, value in enumerate(row, 1):
                        worksheet.cell(row=row_num, column=col_num, value=value)

                # Save the workbook to the specified file path
                workbook.save(file_path)
        except: pass
        self.Key_F3_2_function()
        self.Key_F3_function()

    def clicked_new_book(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, 'Excel fileni tanlang', '', 'Excel Files (*.xlsx *.xls)')

            if file_path:
                # Assuming the Excel file has a sheet named 'Sheet1'
                df = pd.read_excel(file_path, header=None, skiprows=1)

                for index, row in df.iterrows():
                    # Assuming the columns are in order: id, book name, price, barcode, qoldiq
                    book_data = (row[0], row[1], row[2], row[3], row[4], row[6], row[7], str(row[5])[:10])
                    # book_data[-1] = book_data[-1][:4]
                    # book_data[-1] = book_data[-1]
                    # Check if the record already exists based on book name
                    cur.execute("SELECT * FROM Kitob WHERE nomi=?", (book_data[1],))
                    existing_record = cur.fetchone()

                    if existing_record is None:
                        """
                        nomi
                        narxi
                        barcode
                        qoldiq
                        kelgan_sana
                        tan_narx
                        pachka_narx
                        """
                        cur.execute("INSERT INTO Kitob (nomi, narxi, barcode, qoldiq, tan_narx, pachka_narx, kelgan_sana) VALUES (?, ?, ?, ?, ?, ?, ?)", book_data[1:])
                    else:
                        existing_qoldiq = existing_record[4]
                        existing_narxi = existing_record[2]
                        new_qoldiq = int(float(existing_qoldiq)) + int(float(book_data[4]))
                        new_narxi = book_data[2]
                        cur.execute("UPDATE Kitob SET qoldiq=?, narxi=?, nomi=?, kelgan_sana=? WHERE id=?", (new_qoldiq, new_narxi, book_data[1], book_data[-1], book_data[0]))

                conn.commit()
                self.populate_tableWidget_4()
        except: pass
        self.Key_F3_2_function()
        self.Key_F3_function()

    def spacecomma(self, value):
        s_text = ''
        money = str(value)[::-1]
        for i in range(len(money)//3+1):
            s_text+=money[3*i:3*i+3]+' '
        return s_text.strip()[::-1]
        self.Key_F3_function()
    
    def is_numeric(self, value):
        try:
            float(value)
            return True
        except ValueError:
            return False
        
    def populate_tableWidget_4(self, filter=''):
        if not filter: filter = self.lineEdit_3.text()
        self.tableWidget_4.setRowCount(0)
        if filter:
            query = "SELECT id, nomi, tan_narx, narxi, pachka_narx, barcode, qoldiq, kelgan_sana, buyurtma, kimdan FROM Kitob WHERE nomi LIKE ? OR id LIKE ? OR barcode LIKE ? OR kimdan LIKE ? ORDER BY CAST(qoldiq AS SIGNED) DESC"
            cur.execute(query, (f'%{filter}%', f'%{filter}%', f'%{filter}%', f'%{filter}%'))
        else:
            query = "SELECT id, nomi, tan_narx, narxi, pachka_narx, barcode, qoldiq, kelgan_sana, buyurtma, kimdan FROM Kitob ORDER BY CAST(qoldiq AS SIGNED) DESC"
            cur.execute(query)

        data = cur.fetchall()
        num_rows_to_add = 10 + len(data)
        dasmoya = 0
        self.tableWidget_4.setSortingEnabled(False)
        self.tableWidget_4.setRowCount(num_rows_to_add)
        for row_index, row_data in enumerate(data):
            dasmoya += int(float(row_data[2]))*int(float(row_data[6]))
            for col_index, col_data in enumerate(row_data):
                truncated_data = str(col_data)

                if len(truncated_data) > 40:
                    truncated_data = truncated_data[:40] + "..."
                
                item = QTableWidgetItem(truncated_data)
                
                if len(str(col_data)) > 40:
                    item.setToolTip(str(col_data))

                if self.is_numeric(col_data):
                    item.setData(Qt.DisplayRole, int(col_data))
                else:
                    item.setData(Qt.DisplayRole, truncated_data)
                
                self.tableWidget_4.setItem(row_index, col_index, item)
        self.tableWidget_4.setSortingEnabled(True)
        
        money = self.spacecomma(dasmoya)
        self.label_7.setText("Jami dasmoya qiymati: " + money + " so'm")

        for i in range(len(data), num_rows_to_add):
            for j in range(self.tableWidget_4.columnCount()):
                item = QTableWidgetItem("")
                self.tableWidget_4.setItem(i, j, item)
        self.Key_F3_2_function()
        self.Key_F3_function()

    def save_tableWidget_data_to_database(self):
        try:
            for row in range(self.tableWidget_4.rowCount()):
                id = self.tableWidget_4.item(row, 0).text()
                nomi = self.tableWidget_4.item(row, 1).text()
                tan_narx = self.tableWidget_4.item(row, 2).text()
                narxi = self.tableWidget_4.item(row, 3).text()
                pachka_narx = self.tableWidget_4.item(row, 4).text()
                barcode = self.tableWidget_4.item(row, 5).text()
                qoldiq = self.tableWidget_4.item(row, 6).text()
                sana = self.tableWidget_4.item(row, 7).text()
                buyurtma = self.tableWidget_4.item(row, 8).text()
                kimdan = self.tableWidget_4.item(row, 9).text()
                if len(kimdan)==0: kimdan = 'Nomalum'
                if nomi and narxi and barcode and qoldiq:
                    cur.execute("SELECT id FROM Kitob WHERE id=?", (id,))
                    existing_id = cur.fetchone()
                    if existing_id:

                        if float((cur.execute("SELECT qoldiq FROM Kitob WHERE id=?", (id,))).fetchone()[0])==float(qoldiq):
                            # If the record exists, update it
                            cur.execute("""
                                UPDATE Kitob
                                SET nomi=?, narxi=?, barcode=?, qoldiq=?, kelgan_sana=?, tan_narx=?, pachka_narx=?, buyurtma=?, kimdan=?
                                WHERE id=?
                            """, (nomi, int(float(narxi)), barcode, int(float(qoldiq)), sana, int(float(tan_narx)), int(float(pachka_narx)), buyurtma, kimdan, existing_id[0]))
                            conn.commit()
                        else:
                            cur.execute("""
                                UPDATE Kitob
                                SET nomi=?, narxi=?, barcode=?, qoldiq=?, kelgan_sana=?, tan_narx=?, pachka_narx=?, buyurtma=?, kimdan=?
                                WHERE id=?
                            """, (nomi, int(float(narxi)), barcode, int(float(qoldiq)), date.today(), int(float(tan_narx)), int(float(pachka_narx)), buyurtma, kimdan, existing_id[0]))
                            conn.commit()
                    else:
                        """
                        nomi
                        narxi
                        barcode
                        qoldiq
                        kelgan_sana
                        tan_narx
                        pachka_narx
                        """
                        cur.execute("""
                            INSERT INTO Kitob (nomi, tan_narx, narxi, pachka_narx, barcode, qoldiq, kelgan_sana, buyurtma, kimdan)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (nomi, int(float(tan_narx)), int(float(narxi)), int(float(pachka_narx)), barcode, int(float(qoldiq)), date.today(), buyurtma, kimdan))
                        conn.commit()
                elif id and not nomi and not narxi and not barcode and not qoldiq:
                    cur.execute("DELETE FROM Kitob WHERE id=?", (id,))
                    conn.commit()
            QTimer.singleShot(1000, self.close_message_box)
            self.message_info = "Mahsulotlar muaffaqiyatli saqlandi"
            self.message_type = 1
            QTimer.singleShot(50, self.show_message)
            self.populate_tableWidget_4()
            self.Key_F3_2_function()
        except Exception as e:
            print(e)
            QTimer.singleShot(1000, self.close_message_box)
            self.message_info = "Xatolik"
            self.message_type = 0
            QTimer.singleShot(50, self.show_message)
        self.Key_F3_function()

    def show_message(self):
        self.msg_box = QMessageBox()
        self.msg_box.setWindowIcon(QIcon(":/image/logo.jpg"))
        self.msg_box.setWindowTitle('Xabar')
        # icon = QtGui.QIcon()
        # icon.addPixmap(QtGui.QPixmap(":/image/logo.jpg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        # self.msg_box.setIcon(icon)
        self.msg_box.setText(f"{self.message_info}")
        if self.message_type:
            self.msg_box.setStyleSheet("QMessageBox {background-color: lightgray; color: green; font-size: 16px;} QPushButton { color: green; } QLabel {color: green;} QWindowTitle {color: green;}")
        else:
            self.msg_box.setStyleSheet("QMessageBox {background-color: lightgray; color: red; font-size: 16px;} QPushButton { color: red; } QLabel {color: red;} QWindowTitle {color: red;}")
        self.msg_box.exec()
        self.Key_F3_function()

    def close_message_box(self):
        self.msg_box.done(0)
        self.Key_F3_function()

    def sotuv_tarix_tab(self):
        # Date
        today = QDate.currentDate()
        self.dateEdit_2.setDate(today)

        self.dateEdit.setDate(today)
        self.comboBox_3.setCurrentIndex(0)

        self.show_tarix()
        self.filter_history()
        self.Key_F3_function()
    
    def filter_history_combo(self, index):
        today = QDate.currentDate()

        if index == 0:  # "Today"
            self.dateEdit.setDate(today)
            self.dateEdit_2.setDate(today)

        elif index == 1:  # "This week"
            start_of_week = today.addDays(-(today.dayOfWeek() - 1))  # Monday
            end_of_week = start_of_week.addDays(6)  # Sunday
            self.dateEdit.setDate(start_of_week)
            self.dateEdit_2.setDate(end_of_week)

        elif index == 2:  # "A week" (Last 7 days)
            self.dateEdit.setDate(today.addDays(-7))
            self.dateEdit_2.setDate(today)

        elif index == 3:  # "This month"
            start_of_month = QDate(today.year(), today.month(), 1)
            end_of_month = QDate(today.year(), today.month(), start_of_month.daysInMonth())
            self.dateEdit.setDate(start_of_month)
            self.dateEdit_2.setDate(end_of_month)

        elif index == 4:  # "One month" (Last 30 days)
            self.dateEdit.setDate(today.addDays(-30))
            self.dateEdit_2.setDate(today)

        elif index == 5:  # "This year"
            start_of_year = QDate(today.year(), 1, 1)
            end_of_year = QDate(today.year(), 12, 31)
            self.dateEdit.setDate(start_of_year)
            self.dateEdit_2.setDate(end_of_year)

        elif index == 6:  # "One year" (Last 365 days)
            self.dateEdit.setDate(today.addDays(-365))
            self.dateEdit_2.setDate(today)
        self.filter_history()
    
    def filter_history2_combo(self, index):
        today = QDate.currentDate()

        if index == 0:  # "Today"
            self.dateEdit_3.setDate(today)
            self.dateEdit_4.setDate(today)

        elif index == 1:  # "This week"
            start_of_week = today.addDays(-(today.dayOfWeek() - 1))  # Monday
            end_of_week = start_of_week.addDays(6)  # Sunday
            self.dateEdit_3.setDate(start_of_week)
            self.dateEdit_4.setDate(end_of_week)

        elif index == 2:  # "A week" (Last 7 days)
            self.dateEdit_3.setDate(today.addDays(-7))
            self.dateEdit_4.setDate(today)

        elif index == 3:  # "This month"
            start_of_month = QDate(today.year(), today.month(), 1)
            end_of_month = QDate(today.year(), today.month(), start_of_month.daysInMonth())
            self.dateEdit_3.setDate(start_of_month)
            self.dateEdit_4.setDate(end_of_month)

        elif index == 4:  # "One month" (Last 30 days)
            self.dateEdit_3.setDate(today.addDays(-30))
            self.dateEdit_4.setDate(today)

        elif index == 5:  # "This year"
            start_of_year = QDate(today.year(), 1, 1)
            end_of_year = QDate(today.year(), 12, 31)
            self.dateEdit_3.setDate(start_of_year)
            self.dateEdit_4.setDate(end_of_year)

        elif index == 6:  # "One year" (Last 365 days)
            self.dateEdit_3.setDate(today.addDays(-365))
            self.dateEdit_4.setDate(today)
        self.filter_history2()

    def filter_history(self):
        self.tableWidget_3.setRowCount(0)
        start_date = self.dateEdit.date().toString("yyyy-MM-dd")
        end_date = self.dateEdit_2.date().addDays(1).toString("yyyy-MM-dd")

        query = "SELECT * FROM Tarix WHERE sana >= ? AND sana <= ? ORDER BY sana DESC"
        cur.execute(query, (start_date, end_date))
        
        tarix_data = cur.fetchall()
        self.display_data_in_table(tarix_data, self.tableWidget_5)
        query_sum = "SELECT SUM(hisob) AS total_hisob FROM Tarix WHERE sana >= ? AND sana <= ?"
        cur.execute(query_sum, (start_date, end_date))
        total_hisob = cur.fetchone()[0]
        if total_hisob is None:
            total_hisob = 0

        query_payment_types = """
            SELECT tolov_turi, SUM(hisob) 
            FROM Tarix 
            WHERE sana >= ? AND sana <= ? 
            GROUP BY tolov_turi
        """
        cur.execute(query_payment_types, (start_date, end_date))
        payment_type_sums = cur.fetchall()
        payment_summary_text = f"Jami: {self.spacecomma(total_hisob)} so'm\n"
        for payment_type, total in payment_type_sums:
            payment_summary_text += f"{payment_type}: {self.spacecomma(total)} so'm\n"
        self.label_8.setText(payment_summary_text)
        self.Key_F3_function()

    def handle_tableWidget_5_selected(self):
        self.tableWidget_3.setRowCount(0)
        selected_items = self.tableWidget_5.selectedItems()
        row = 0
        for item in selected_items: row = int(float(item.row()))
        if selected_items: self.show_table_widget_3(row)
        self.Key_F3_function()
    
    def show_table_widget_3(self, row):
        try:
            tarix_id = int(float(self.tableWidget_5.item(row, 0).text()))
            cur.execute("""
                SELECT TarixItem.id, Kitob.nomi, TarixItem.soni, TarixItem.hisob
                FROM TarixItem
                JOIN Kitob ON TarixItem.Kitob = Kitob.id
                WHERE TarixItem.Tarix = ?
            """, (tarix_id,))
            data = cur.fetchall()
            self.display_data_in_table(data, self.tableWidget_3)
        except:
            pass
        self.Key_F3_function()

    def show_tarix(self):
        cur.execute("SELECT * FROM Tarix ORDER BY sana DESC")
        tarix_data = cur.fetchall()
        self.display_data_in_table(tarix_data, self.tableWidget_5)
        cur.execute("SELECT SUM(hisob) FROM Tarix ORDER BY sana DESC")
        total_hisob = cur.fetchone()[0]
        if total_hisob is None: total_hisob=0
        self.label_8.setText(f"Jami: {self.spacecomma(total_hisob)} so'm")
        self.Key_F3_function()
    
    def display_data_in_table(self, data, table_widget):
        try:
            table_widget.setRowCount(len(data))
            table_widget.setColumnCount(len(data[0]))

            for row_index, row_data in enumerate(data):
                for col_index, col_data in enumerate(row_data):
                    truncated_data = str(col_data)

                    if len(truncated_data) > 40:
                        truncated_data = truncated_data[:40] + "..."
                    item = QTableWidgetItem(truncated_data)
                    
                    if len(str(col_data)) > 40:
                        item.setToolTip(str(col_data))

                    table_widget.setItem(row_index, col_index, item)
        except:
            pass
        self.Key_F3_function()

    def tovar_tab(self):
        self.Key_F3_function2()

        self.timer = QTimer(self)
        self.timer.setSingleShot(True)
        self.timer.timeout.connect(self.search_database2)
        self.Key_F3_function()

    def store_new_product(self, p):
        cur.execute("INSERT INTO Kitob (nomi, tan_narx, narxi, pachka_narx, barcode, qoldiq, kelgan_sana, kimdan) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", 
                    (p[0], p[1], p[2], p[3], p[4], p[5], p[6], p[7]))
        conn.commit()
        return cur.lastrowid

    def Key_F3_function2(self):
        self.lineEdit_4.setFocus()

    def scanner_returned2(self):
        scanned_text = self.lineEdit_4.text()
        self.Key_F3_function()

    def clicked_cancel2(self, item):
        if item.column() == 9:
            if 0 <= item.row() < self.tableWidget_7.rowCount():
                self.tableWidget_7.removeRow(item.row())
            else:
                pass
            self.calculate2()
        self.Key_F3_function()

    def handle_item_changed2(self, item):
        # Check if the item is in column 2
        if item.column() in [2, 3]:
            try: self.calculate2()
            except: self.cancel_buy2()
        self.Key_F3_function()

    def calculate2(self):
        column_index = 2
        total_amount = 0

        for row_index in range(self.tableWidget_7.rowCount()):
            item = self.tableWidget_7.item(row_index, column_index)
            item2 = self.tableWidget_7.item(row_index, 3)

            if item is not None and item2 is not None:
                add = int(float(item.text()))*int(float(item2.text()))
                total_amount+=add
                self.tableWidget_7.setItem(row_index, 8, QTableWidgetItem(str(add)))

        self.lineEdit_5.setText(self.spacecomma(int(total_amount)))
        self.Key_F3_function2()
        self.Key_F3_function()

    def cancel_buy2(self):
        self.tableWidget_6.setRowCount(0)
        self.tableWidget_7.setRowCount(0)
        self.lineEdit_5.clear()
        self.lineEdit_6.clear()
        self.lineEdit_5.setText("0")
        self.Key_F3_function2()
        self.Key_F3_function()

    def accept_buy2(self):
        try:
            rows = self.tableWidget_7.rowCount()
            if rows:
                current_datetime = datetime.now()
                formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
                tex  = self.lineEdit_5.text()
                umumiy_hisob = ''
                for i in tex:
                    if i.isdigit():
                        umumiy_hisob+=i
                kimdan = self.lineEdit_6.text()
                if len(kimdan) < 3:
                    QTimer.singleShot(5000, self.close_message_box)
                    self.message_info = "Kiritish xatosi!\nTovarlarni kimdan olganingizni kiriting!\nEng kichik ism uzunligi 3 ta belgi bo'lishi mumkin!"
                    self.message_type = 0
                    QTimer.singleShot(50, self.show_message)
                    return
                cur.execute("INSERT INTO Tovar (kimdan, sana, hisob) VALUES (?, ?, ?)", (kimdan, formatted_datetime, umumiy_hisob))
                conn.commit()
                tarix_id = cur.lastrowid
                for row_index in range(rows):
                    nomi = str(self.tableWidget_7.item(row_index, 1).text())
                    soni = int(float(self.tableWidget_7.item(row_index, 2).text()))
                    tan_narx = int(float(self.tableWidget_7.item(row_index, 3).text()))
                    narx = int(float(self.tableWidget_7.item(row_index, 4).text()))
                    pachka_narx = int(float(self.tableWidget_7.item(row_index, 5).text()))
                    barcode = self.tableWidget_7.item(row_index, 6).text()
                    hisob = int(float(self.tableWidget_7.item(row_index, 8).text()))
                    try:
                        kitob_id = int(float(self.tableWidget_7.item(row_index, 0).text()))
                        cur.execute("SELECT qoldiq FROM Kitob where id = ?", (kitob_id,))
                        qoldiq = int(cur.fetchone()[0])
                        cur.execute("UPDATE Kitob SET nomi = ?, tan_narx = ?, narxi = ?, pachka_narx = ?, barcode = ?, qoldiq = ?, kelgan_sana=?, kimdan=? WHERE id = ?", (nomi, tan_narx, narx, pachka_narx, barcode, qoldiq+soni, date.today(), kimdan, kitob_id))
                        conn.commit()
                    except Exception as e:
                        qoldiq = 0
                        product = (nomi, tan_narx, narx, pachka_narx, barcode, qoldiq+soni, date.today(), kimdan)
                        kitob_id = self.store_new_product(product)
                        
                    cur.execute("INSERT INTO TovarItem (Tovar, Kitob, soni, hisob, tan_narx, sotuv_narx) VALUES (?, ?, ?, ?, ?, ?)", (tarix_id, kitob_id, soni, hisob, tan_narx, narx))
                    conn.commit()

            QTimer.singleShot(1000, self.close_message_box)
            self.message_info = "Saqlandi"
            self.message_type = 1
            QTimer.singleShot(50, self.show_message)
        except Exception as e:
            print(e)
            QTimer.singleShot(1000, self.close_message_box)
            self.message_info = "Xatolik"
            self.message_type = 0
            QTimer.singleShot(50, self.show_message)
        self.cancel_buy2()
        self.Key_F3_function2()
        self.Key_F3_function()
    
    def on_line_edit_changed2(self):
        # Start the timer when the text in the line edit changes
        self.timer.start(1000)  # 1000 milliseconds = 1 second
        self.Key_F3_function()

    def search_database2(self):
        # Perform the database query based on user input
        user_input = self.lineEdit_4.text()
        try:
            if user_input:
                if len(user_input)>7 and user_input.isdigit():
                    query = "SELECT id, nomi, narxi, qoldiq, barcode FROM Kitob WHERE (barcode = ?)"
                    cur.execute(query, (user_input,))
                    result = cur.fetchall()
                    if len(result)==1:
                        self.display_data_in_table2(result, self.tableWidget_6)
                        self.tableWidget_6.selectRow(0)
                        self.add_selected_item_to_table2()
                    elif len(result)>1:
                        self.tableWidget_6.selectRow(0)
                        self.display_data_in_table2(result, self.tableWidget_6)
                    else:
                        row_position = self.tableWidget_7.rowCount()
                        self.tableWidget_7.insertRow(row_position)
                        self.tableWidget_7.setItem(row_position, 6, QtWidgets.QTableWidgetItem(str(user_input)))
                        self.tableWidget_7.setItem(row_position, 2, QtWidgets.QTableWidgetItem(str(1)))
                        self.tableWidget_7.setItem(row_position, 3, QtWidgets.QTableWidgetItem(str(0)))
                        self.tableWidget_7.setItem(row_position, 4, QtWidgets.QTableWidgetItem(str(0)))
                        self.tableWidget_7.setItem(row_position, 5, QtWidgets.QTableWidgetItem(str(0)))
                        self.tableWidget_7.setItem(row_position, 7, QtWidgets.QTableWidgetItem(str(0)))
                        self.tableWidget_7.setItem(row_position, 8, QtWidgets.QTableWidgetItem(str(0)))
                        column_position = 9
                        item = QTableWidgetItem('Bekor qilish')
                        background_color = QColor(160, 20, 20)
                        color = QColor(255, 255, 255)
                        item.setBackground(background_color)
                        item.setForeground(color)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        self.tableWidget_7.setItem(row_position, column_position, item)
                else:
                    query = "SELECT id, nomi, narxi, qoldiq, barcode FROM Kitob WHERE (nomi LIKE ? OR id LIKE ?)"
                    cur.execute(query, (f'%{user_input}%', f'%{user_input}%'))
                    result = cur.fetchall()
                    if len(result):
                        self.tableWidget_6.selectRow(0)
                        self.display_data_in_table2(result, self.tableWidget_6)
        except Exception as e:
            print(e)
        self.Key_F3_function()

    def add_selected_item_to_table2(self):
        selected_item = self.tableWidget_6.currentRow()
        self.lineEdit_4.clear()
        if selected_item>-1:
            id = int(float(self.tableWidget_6.item(selected_item, 0).text()))
            check = self.check_tablewidget_add2(id)
            if check==-1:
                row_position = self.tableWidget_7.rowCount()
                self.tableWidget_7.insertRow(row_position)
                cur.execute("SELECT id, nomi, tan_narx, narxi, pachka_narx, barcode, qoldiq FROM Kitob WHERE id=?", (id,))
                data = cur.fetchone()
                # if int(float(data[4])):
                for col_num, col_data in enumerate(data):
                    item = QTableWidgetItem(str(col_data))
                    if col_num>1: col_num+=1
                    self.tableWidget_7.setItem(row_position, col_num, item)

                self.tableWidget_7.setItem(row_position, 2, QTableWidgetItem('1'))
                column_position = 9
                item = QTableWidgetItem('Bekor qilish')
                background_color = QColor(160, 20, 20)
                color = QColor(255, 255, 255)
                item.setBackground(background_color)
                item.setForeground(color)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.tableWidget_7.setItem(row_position, column_position, item)
                # item = self.tableWidget_7.item(row_position, 2)
                # item.setBackground(QColor(0, 0, 255))
                # item.setForeground(QColor(255, 255, 255))
                # item = self.tableWidget_7.item(row_position, 3)
                # item.setBackground(QColor(60, 179, 113))
                # item.setForeground(QColor(255, 255, 255))
            else:
                prev = int(float(self.tableWidget_7.item(check, 2).text()))
                self.tableWidget_7.setItem(check, 2, QTableWidgetItem(str(prev+1)))
                # item = self.tableWidget_7.item(check, 2)
                # item.setBackground(QColor(0, 0, 255))
                # item.setForeground(QColor(255, 255, 255))
        # self.tableWidget_6.setRowCount(0)
        self.Key_F3_function2()
        self.Key_F3_function()

    def subtract_selected_item_to_table2(self):
        selected_item = self.tableWidget_6.currentRow()
        self.lineEdit_4.clear()
        if selected_item>-1:
            id = int(float(self.tableWidget_6.item(selected_item, 0).text()))
            check = self.check_tablewidget_add2(id)
            if check==-1:
                row_position = self.tableWidget_7.rowCount()
                self.tableWidget_7.insertRow(row_position)
                cur.execute("SELECT id, nomi, tan_narx, narxi, pachka_narx, barcode, qoldiq FROM Kitob WHERE id=?", (id,))
                data = cur.fetchone()
                # if int(float(data[4])):
                for col_num, col_data in enumerate(data):
                    item = QTableWidgetItem(str(col_data))
                    if col_num>1: col_num = max(0, col_num - 1)
                    self.tableWidget_7.setItem(row_position, col_num, item)

                self.tableWidget_7.setItem(row_position, 2, QTableWidgetItem('1'))
                column_position = 9
                item = QTableWidgetItem('Bekor qilish')
                background_color = QColor(160, 20, 20)
                color = QColor(255, 255, 255)
                item.setBackground(background_color)
                item.setForeground(color)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.tableWidget_7.setItem(row_position, column_position, item)
                # item = self.tableWidget_7.item(row_position, 2)
                # item.setBackground(QColor(0, 0, 255))
                # item.setForeground(QColor(255, 255, 255))
                # item = self.tableWidget_7.item(row_position, 3)
                # item.setBackground(QColor(60, 179, 113))
                # item.setForeground(QColor(255, 255, 255))
            else:
                prev = int(float(self.tableWidget_7.item(check, 2).text()))
                self.tableWidget_7.setItem(check, 2, QTableWidgetItem(str(max(0, prev-1))))
                # item = self.tableWidget_7.item(check, 2)
                # item.setBackground(QColor(0, 0, 255))
                # item.setForeground(QColor(255, 255, 255))
        # self.tableWidget_6.setRowCount(0)
        self.Key_F3_function2()
        self.Key_F3_function()

    def check_tablewidget_add2(self, id):
        is_present = -1
        for row_index in range(self.tableWidget_7.rowCount()):
            try:
                if self.tableWidget_7.item(row_index, 0).text() == str(id):
                    is_present=row_index
                    break
            except:
                pass
        return is_present

    def tovar_tarix_tab(self):
        today = QDate.currentDate()
        self.dateEdit_4.setDate(today)

        # Calculate the date 7 days back from today
        self.dateEdit_3.setDate(today)
        self.comboBox_4.setCurrentIndex(0)

        self.show_tarix2()
        self.filter_history2()
        self.Key_F3_function()
    
    def filter_history2(self):
        self.tableWidget_9.setRowCount(0)
        start_date = self.dateEdit_3.date().toString("yyyy-MM-dd")
        end_date = self.dateEdit_4.date().addDays(1).toString("yyyy-MM-dd")

        query = "SELECT * FROM Tovar WHERE sana >= ? AND sana <= ? ORDER BY sana DESC"
        cur.execute(query, (start_date, end_date))
        
        tarix_data = cur.fetchall()
        self.display_data_in_table2(tarix_data, self.tableWidget_8)
        query_sum = "SELECT SUM(hisob) AS total_hisob FROM Tovar WHERE sana >= ? AND sana <= ?"
        cur.execute(query_sum, (start_date, end_date))  # Replace start_date and end_date with your actual date values
        total_hisob = cur.fetchone()[0]
        if total_hisob is None: total_hisob=0
        self.label_15.setText(f"Jami: {self.spacecomma(total_hisob)} so'm")
        self.Key_F3_function()

    def handle_tableWidget_8_selected(self):
        self.tableWidget_9.setRowCount(0)
        selected_items = self.tableWidget_8.selectedItems()
        row = 0
        for item in selected_items: row = int(float(item.row()))
        if selected_items: self.show_table_widget_8_tovar(row)
        self.Key_F3_function()
    
    def show_table_widget_8_tovar(self, row):
        try:
            tarix_id = int(float(self.tableWidget_8.item(row, 0).text()))
            cur.execute("""
                SELECT TovarItem.id, Kitob.nomi, TovarItem.soni, TovarItem.hisob, TovarItem.tan_narx, TovarItem.sotuv_narx
                FROM TovarItem
                JOIN Kitob ON TovarItem.Kitob = Kitob.id
                WHERE TovarItem.Tovar = ?
            """, (tarix_id,))
            data = cur.fetchall()
            self.display_data_in_table2(data, self.tableWidget_9)
        except:
            pass
        self.Key_F3_function()

    def show_tarix2(self):
        cur.execute("SELECT * FROM Tovar ORDER BY sana DESC")
        tarix_data = cur.fetchall()
        self.display_data_in_table2(tarix_data, self.tableWidget_8)
        cur.execute("SELECT SUM(hisob) FROM Tovar ORDER BY sana DESC")
        total_hisob = cur.fetchone()[0]
        if total_hisob is None: total_hisob=0
        self.label_15.setText(f"Jami: {self.spacecomma(total_hisob)} so'm")
        self.Key_F3_function()
    
    def display_data_in_table2(self, data, table_widget):
        try:
            table_widget.setRowCount(len(data))
            table_widget.setColumnCount(len(data[0]))

            for row_index, row_data in enumerate(data):
                for col_index, col_data in enumerate(row_data):
                    truncated_data = str(col_data)

                    if len(truncated_data) > 40:
                        truncated_data = truncated_data[:40] + "..."
                    
                    item = QTableWidgetItem(truncated_data)
                    if len(str(col_data)) > 40:
                        item.setToolTip(str(col_data))

                    table_widget.setItem(row_index, col_index, item)
        except:
            pass
        self.Key_F3_function()

if __name__ == "__main__":
    App = QApplication(sys.argv)
    wind = DiamondWindow()
    wind.show()
    sys.exit(App.exec())
